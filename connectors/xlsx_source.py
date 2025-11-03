# connectors/xlsx_source.py
from .base import DataSource
from typing import Any, Dict, List
import openpyxl
from openpyxl.utils import get_column_letter
import os
import logging

"""не определяет автоматом ключевые поля"""

# Настройка логирования для этого модуля
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)  # Создаем логгер для этого файла

class XlsxDataSource(DataSource):
    def __init__(self, connection_params: Dict[str, Any]):
        # Ожидаем, что connection_params будет содержать 'path' - путь к .xlsx файлу
        self.file_path = connection_params.get("path")
        if not self.file_path:
            error_msg = "Connection parameter 'path' is required for Excel file."
            logger.error(error_msg)
            raise ValueError(error_msg)
        if not self.file_path.lower().endswith('.xlsx'):
            error_msg = "File must have .xlsx extension."
            logger.error(error_msg)
            raise ValueError(error_msg)

        # Сохраняем путь, но соединение устанавливается при вызове connect()
        self._is_read_only = connection_params.get("read_only", False)  # Добавим флаг для режима чтения
        logger.debug(f"Initialized XlsxDataSource with file_path: {self.file_path}, read_only: {self._is_read_only}")

    def connect(self):
        try:
            logger.info(f"Attempting to load Excel workbook at: {self.file_path} in read_only={self._is_read_only} mode.")
            if self._is_read_only:
                # load_workbook в режиме read_only быстрее и потребляет меньше памяти
                self.workbook = openpyxl.load_workbook(self.file_path, read_only=True)
            else:
                # Для записи/редактирования
                self.workbook = openpyxl.load_workbook(self.file_path)
            logger.info("Excel workbook loaded successfully.")
        except FileNotFoundError:
            error_msg = f"Excel file not found: {self.file_path}"
            logger.error(error_msg)
            raise Exception(error_msg)
        except Exception as err:
            logger.error(f"Excel connection failed: {err}")
            raise Exception(f"Excel connection failed: {err}")

    def fetch_data(self, query: str, params: Dict[str, Any] = None) -> List[Dict[str, Any]]:
        """
        Извлекает данные из листа Excel.
        query: имя листа (например, 'Sheet1')
        params: не используется для openpyxl
        """
        if not self.workbook:
            error_msg = "Workbook not loaded. Call connect() first."
            logger.error(error_msg)
            raise Exception(error_msg)

        if params:
            logger.warning("Warning: params are not used in XlsxDataSource.fetch_data")

        try:
            logger.debug(f"Fetching data from worksheet: {query}")
            worksheet = self.workbook[query]
        except KeyError:
            error_msg = f"Worksheet '{query}' not found in the workbook."
            logger.error(error_msg)
            raise Exception(error_msg)

        rows = worksheet.iter_rows(values_only=True)
        data = list(rows)

        if not data:  # ОШИБКА: отсутствовало условие
            logger.debug("Worksheet is empty, returning empty list.")
            return []

        # Предполагаем, что первая строка - заголовки
        headers = data[0]
        logger.debug(f"Headers found: {headers}")
        # Пропускаем заголовки и формируем словари
        result = []
        for i, row in enumerate(data[1:], start=2): # Начинаем с индекса 2 для логирования номера строки
            # Пропускаем пустые строки (если все значения None)
            if any(cell is not None for cell in row):
                result.append(dict(zip(headers, row)))
            else:
                logger.debug(f"Skipping empty row at index {i}.")
        logger.info(f"Fetched {len(result)} rows from worksheet '{query}'.")
        return result

    def build_select_query(self, table_name: str, fields: List[str] = None) -> str:
        """
        Для XlsxDataSource, table_name - это имя листа.
        Возвращаем просто имя листа, так как SQL-запросы не используются.
        """
        # fields игнорируются, так как openpyxl не использует SQL.
        # В fetch_data нужно будет отфильтровать вручную, если нужно.
        logger.debug(f"Building select query for worksheet: {table_name}")
        return table_name

    def get_schema(self) -> Dict[str, Any]:
        """
        Возвращает информацию о листах и первых строках (предполагаем заголовки).
        """
        if not self.workbook:
            error_msg = "Workbook not loaded. Call connect() first."
            logger.error(error_msg)
            raise Exception(error_msg)

        try:
            logger.info("Fetching workbook schema.")
            schema = {}
            for sheet_name in self.workbook.sheetnames:
                worksheet = self.workbook[sheet_name]

                # Получаем первую строку (заголовки)
                headers = []
                logger.debug(f"Reading headers from worksheet '{sheet_name}'.")
                for col_idx, cell in enumerate(worksheet[1], start=1):  # worksheet[1] - первая строка
                    # Имя колонки может быть None, если ячейка пуста
                    header_name = cell.value if cell.value is not None else f"Column_{get_column_letter(col_idx)}"
                    headers.append(header_name)

                # Предполагаем, что тип данных - это тип значения в первой строке данных (если есть)
                # Это грубое приближение, т.к. типы могут меняться в столбце
                schema[sheet_name] = []
                logger.debug(f"Headers for '{sheet_name}': {headers}")
                for header in headers:
                    # Так как Excel не хранит типы строго, мы можем только зафиксировать имя
                    # и, возможно, тип первого значения в данных (строка 2)
                    # Пока оставим тип как строку, можно улучшить
                    schema[sheet_name].append({
                        "name": header,
                        "type": "unknown",  # openpyxl не предоставляет строгий тип колонки
                        "not_null": False,  # Excel не имеет ограничений NOT NULL
                        "default": None,
                        "extra": "",
                        "primary_key": False  # Excel не имеет первичных ключей
                    })
            logger.info(f"Schema fetched for {len(schema)} worksheets.")
            return schema
        except Exception as e:
            logger.error(f"An error occurred during get_schema: {e}")
            raise

    # --- Новые методы для INSERT, UPDATE, DELETE ---
    # ВНИМАНИЕ: Эти операции требуют открытия файла в режиме записи (read_only=False).
    # Их выполнение может быть сложнее, чем в базах данных.
    # Также файл не должен быть открыт вручную в Excel во время операций.
    def insert_data(self, data: List[Dict[str, Any]], table_name: str):
        """
        Вставляет новые строки в лист Excel.
        Если лист с именем table_name не существует, он будет создан.
        data: список словарей с данными. Ключи должны соответствовать заголовкам в Excel.
        table_name: имя листа.
        """
        if self._is_read_only:
            error_msg = "Cannot insert data: workbook opened in read-only mode."
            logger.error(error_msg)
            raise Exception(error_msg)

        if not self.workbook:
            error_msg = "Workbook not loaded. Call connect() first."
            logger.error(error_msg)
            raise Exception(error_msg)

        if not data:
            logger.info("No data provided for insertion, skipping.")
            return

        # Убедиться, что лист существует
        self.ensure_worksheet_exists(table_name)

        worksheet = self.workbook[table_name]  # Теперь этот вызов не должен вызывать KeyError


        # Получаем заголовки из Excel (предполагаем, что они в строке 1)
        # Если лист новый, заголовки будут пустыми
        headers = [cell.value for cell in worksheet[1]]
        logger.debug(f"Headers in worksheet '{table_name}': {headers}")

        # Если заголовки пусты (новый лист), устанавливаем их из ключей первого элемента данных
        if not headers or all(h is None for h in headers):
            if data:
                headers = list(data[0].keys())
                for col_idx, header in enumerate(headers, start=1):
                    worksheet.cell(row=1, column=col_idx, value=header)
                logger.info(f"Set headers for new worksheet '{table_name}': {headers}")
            else:
                logger.info(f"Worksheet '{table_name}' is empty and no data to insert, headers remain empty.")
                return  # Нечего вставлять и нечего устанавливать как заголовки

        # Проверяем, что все ключи в data присутствуют в заголовках
        for i, row_data in enumerate(data):
            for key in row_data.keys():
                if key not in headers:
                    error_msg = f"Column '{key}' not found in worksheet '{table_name}' headers at row {i + 1} (data: {row_data}). Expected headers: {headers}"
                    logger.error(error_msg)
                    raise ValueError(error_msg)

        # Найти следующую строку для вставки (после заголовков или после последней заполненной)
        # Если лист был пуст или только что создан с заголовками, max_row будет 1
        next_row_idx = worksheet.max_row + 1 if worksheet.max_row > 1 or any(
            cell.value is not None for cell in worksheet[1]) else 2
        logger.debug(f"Inserting rows starting from row index: {next_row_idx}")

        for i, row_data in enumerate(data):
            # Создаем список значений в том же порядке, что и заголовки
            row_values = [row_data.get(header, None) for header in headers]
            for col_idx, value in enumerate(row_values, start=1):
                worksheet.cell(row=next_row_idx + i, column=col_idx, value=value)
            logger.debug(f"Inserted row {i + 1}: {row_data}")

        logger.info(f"Successfully inserted {len(data)} rows into worksheet '{table_name}'.")

    def update_data(self, updates: List[Dict[str, Any]], key_fields: List[str], table_name: str):
        """
        Обновляет строки в листе Excel на основе ключевых полей.
        Если лист не существует, бросает ошибку.
        """
        if self._is_read_only:
            error_msg = "Cannot update data: workbook opened in read-only mode."
            logger.error(error_msg)
            raise Exception(error_msg)

        if not self.workbook:
            error_msg = "Workbook not loaded. Call connect() first."
            logger.error(error_msg)
            raise Exception(error_msg)

        if not updates:
            logger.info("No updates provided, skipping.")
            return  # Нечего обновлять

        try:
            logger.info(f"Attempting to update rows in worksheet '{table_name}' using key fields: {key_fields}.")
            worksheet = self.workbook[table_name]
        except KeyError:
            error_msg = f"Worksheet '{table_name}' not found in the workbook for update operation."
            logger.error(error_msg)
            raise Exception(error_msg)

        headers = [cell.value for cell in worksheet[1]]
        logger.debug(f"Headers in worksheet '{table_name}': {headers}")

        if not headers or all(h is None for h in headers):
            logger.warning(f"Worksheet '{table_name}' has no headers or empty headers. Cannot update data.")
            return  # Нечего обновлять

        for kf in key_fields:
            if kf not in headers:
                error_msg = f"Key field '{kf}' not found in worksheet '{table_name}' headers."
                logger.error(error_msg)
                raise ValueError(error_msg)

        updated_count = 0
        for row_idx in range(2, worksheet.max_row + 1):
            current_row_dict = {}
            for col_idx, header in enumerate(headers, start=1):
                current_row_dict[header] = worksheet.cell(row=row_idx, column=col_idx).value

            for update_item in updates:
                old_row = update_item["old"]
                new_row = update_item["new"]

                match = all(current_row_dict.get(kf) == old_row.get(kf) for kf in key_fields)
                if match:
                    logger.debug(
                        f"Match found at row {row_idx} for keys {key_fields} with old values {old_row}. Updating to {new_row}.")
                    for key, value in new_row.items():
                        if key in headers:
                            col_idx = headers.index(key) + 1
                            worksheet.cell(row=row_idx, column=col_idx, value=value)
                    updated_count += 1
                    break

        logger.info(f"Successfully updated {updated_count} rows in worksheet '{table_name}'.")

    def delete_data(self, deletions: List[Dict[str, Any]], key_fields: List[str], table_name: str):
        """
        Удаляет строки из листа Excel на основе ключевых полей.
        Если лист не существует, бросает ошибку.
        """
        if self._is_read_only:
            error_msg = "Cannot delete data: workbook opened in read-only mode."
            logger.error(error_msg)
            raise Exception(error_msg)

        if not self.workbook:
            error_msg = "Workbook not loaded. Call connect() first."
            logger.error(error_msg)
            raise Exception(error_msg)

        if not deletions:
            logger.info("No deletions provided, skipping.")
            return

        try:
            logger.info(f"Attempting to delete rows from worksheet '{table_name}' using key fields: {key_fields}.")
            worksheet = self.workbook[table_name]
        except KeyError:
            error_msg = f"Worksheet '{table_name}' not found in the workbook for delete operation."
            logger.error(error_msg)
            raise Exception(error_msg)

        headers = [cell.value for cell in worksheet[1]]
        logger.debug(f"Headers in worksheet '{table_name}': {headers}")

        if not headers or all(h is None for h in headers):
            logger.warning(f"Worksheet '{table_name}' has no headers or empty headers. Cannot delete data.")
            return  # Нечего удалять

        for kf in key_fields:
            if kf not in headers:
                error_msg = f"Key field '{kf}' not found in worksheet '{table_name}' headers."
                logger.error(error_msg)
                raise ValueError(error_msg)

        rows_to_delete = []
        for row_idx in range(2, worksheet.max_row + 1):
            current_row_dict = {}
            for col_idx, header in enumerate(headers, start=1):
                current_row_dict[header] = worksheet.cell(row=row_idx, column=col_idx).value

            for del_row in deletions:
                match = all(current_row_dict.get(kf) == del_row.get(kf) for kf in key_fields)
                if match:
                    logger.debug(
                        f"Match found at row {row_idx} for deletion with keys {key_fields} and values {del_row}.")
                    rows_to_delete.append(row_idx)
                    break

        rows_to_delete.sort(reverse=True)
        logger.debug(f"Rows marked for deletion (in reverse order): {rows_to_delete}")

        for row_idx in rows_to_delete:
            worksheet.delete_rows(row_idx, 1)
            logger.debug(f"Deleted row at index {row_idx}.")

        logger.info(f"Successfully deleted {len(rows_to_delete)} rows from worksheet '{table_name}'.")

    def disconnect(self):
        try:
            if self.workbook:
                # Если файл открывался в режиме записи, нужно его сохранить
                if not self._is_read_only:
                    logger.info(f"Saving workbook to {self.file_path}")
                    try:
                        self.workbook.save(self.file_path)
                        logger.info(f"Workbook saved to {self.file_path}")
                    except PermissionError:
                        error_msg = f"Could not save file: {self.file_path} might be open in Excel or lack write permissions."
                        logger.error(error_msg)
                        raise Exception(error_msg)
                else:
                    logger.info("Workbook was opened in read-only mode, skipping save.")
                self.workbook.close()
                logger.info("Excel workbook closed.")
                self.workbook = None
        except Exception as e:
            logger.error(f"An error occurred during disconnect: {e}")
            raise # Возбуждаем исключение дальше, если это критично

    def standard_formatting(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Стандартное форматирование данных после выборки из Excel.
        openpyxl обычно возвращает правильные типы Python (str, int, float, datetime, bool, None).
        """
        # logger.debug(f"Standard formatting completed for {len(data)} rows (XlsxDataSource typically requires no additional formatting).")
        return data

    def ensure_worksheet_exists(self, table_name: str):
        """Создаёт лист с именем table_name, если он не существует."""
        if table_name not in self.workbook.sheetnames:
            logger.info(f"Worksheet '{table_name}' does not exist. Creating it.")
            self.workbook.create_sheet(title=table_name)
        else:
            logger.debug(f"Worksheet '{table_name}' already exists.")